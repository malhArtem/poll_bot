import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from db.models.tests import Tests
from db.models.tests_answers import Answers
from db.models.tests_questions import Questions


async def parce_excel(filename, session):
    book = openpyxl.open(filename)
    sheet: Worksheet = book.active
    print(session)
    settings = parce_settings(sheet)
    test = await Tests.add_test(**settings, session=session)
    print(session)
    print(test.id)
    if test is None:
        return None

    for row in range(14, sheet.max_row, 3):
        if not sheet.cell(row=row, column=2).value:
            break

        question = parce_question(sheet, row)
        question = await Questions.add_question(test_id=test.id, session=session, **question)
        if question is None:
            return None

        for col in range(4, 8):
            if not sheet.cell(row=row, column=col):
                break

            answer = parce_answer(sheet, row, col)
            answer = await Answers.add_answer(test_id=test.id, session=session, question_id=question.id, **answer)
            if answer is None:
                return None

    return test


def parce_settings(sheet: Worksheet) -> dict:
    question = {}
    question['name'] = sheet.cell(row=1, column=3).value
    question['is_free'] = True if sheet.cell(row=2, column=3).value.lower() == "да" else False
    question['show_answer'] = True if sheet.cell(row=3, column=3).value.lower() == "да" else False
    question['invite_answer'] = True if sheet.cell(row=4, column=3).value.lower() == "да" else False
    question['notify'] = True if sheet.cell(row=5, column=3).value.lower() == "да" else False
    question['detailed_analysis'] = sheet.cell(row=6, column=3).value

    question['first_message'] = sheet.cell(row=12, column=3).value
    question['last_message'] = sheet.cell(row=13, column=3).value
    question['points_greed'] = ''
    for i in range(7, 12):
        if sheet.cell(row=i, column=3).value and sheet.cell(row=i, column=4).value:
            question['points_greed'] += f"{sheet.cell(row=i, column=3).value}: {sheet.cell(row=i, column=4)}"

    return question


def parce_question(sheet: Worksheet, row) -> dict:
    question = {}
    question['question_number'] = int(sheet.cell(row, 2).value)
    question['question'] = sheet.cell(row, 3).value
    return question


def parce_answer(sheet: Worksheet, row, col) -> dict:
    answer = {}
    answer['answer'] = sheet.cell(row, col).value
    answer['comment'] = sheet.cell(row + 1, col).value
    answer['score'] = int(sheet.cell(row + 2, col).value)
    return answer

